from calendar import monthrange
import pandas as pds
import openpyxl
from openpyxl.styles import PatternFill

# check = monthrange(2021, 3)
excelName = "reporddst.xlsx"

wb = openpyxl.load_workbook("C:\\Users\\S.S\\PycharmProjects\\locationData\\test.xlsx")
sh1 = wb["Sheet1"]

newData = pds.read_excel(r'C:\Users\S.S\PycharmProjects\locationData\Insurance.xlsx', sheet_name="Sheet1")


def fetchData(start, end):
    delhiCount = 0
    nodiaCount = 0
    gazibadCount = 0
    gurgoanCount = 0
    faridadCount = 0
    chatishgarhCount = 0
    jaipurCount = 0
    mumbaiCount = 0
    puneCount = 0
    ahmedabadCount = 0
    nagarCount = 0
    suratCount = 0
    benguluruCount = 0
    hyderabadCount = 0
    chennaiCount = 0
    kochiCount = 0
    coimbatoreCount = 0
    kanpurCount = 0
    lucknowCount = 0
    indoreCount = 0
    kolkataCount = 0

    findMasterLoop = end[1] - start[1]
    masterLoop = 0
    month = 0
    loopvalue = 0
    noOfMonth = []

    sh1.cell(2, column=1, value="Delhi")
    sh1.cell(3, column=1, value="Noida")
    sh1.cell(4, column=1, value="Ghaziabad")
    sh1.cell(5, column=1, value="Gurgaon")
    sh1.cell(6, column=1, value="Faridabad")
    sh1.cell(7, column=1, value="Chandigarh")
    sh1.cell(8, column=1, value="Jaipur")
    sh1.cell(9, column=1).fill = PatternFill("solid", fgColor="FFFF00")
    sh1.cell(9, column=1, value="Delhi NCR")
    sh1.cell(10, column=1, value="Mumbai")
    sh1.cell(11, column=1, value="Pune")
    sh1.cell(12, column=1, value="Ahmedabad")
    sh1.cell(13, column=1, value="Nagpur")
    sh1.cell(14, column=1, value="Surat")
    sh1.cell(15, column=1).fill = PatternFill("solid", fgColor="FFFF00")
    sh1.cell(15, column=1, value="West")
    sh1.cell(16, column=1, value="Bengaluru")
    sh1.cell(17, column=1, value="Hyderabad")
    sh1.cell(18, column=1, value="Chennai")
    sh1.cell(19, column=1).fill = PatternFill("solid", fgColor="FFFF00")
    sh1.cell(19, column=1, value="South")
    sh1.cell(20, column=1, value="Kochi")
    sh1.cell(21, column=1, value="Coimbatore")
    sh1.cell(22, column=1, value="Kanpur")
    sh1.cell(23, column=1, value="Lucknow")
    sh1.cell(24, column=1, value="Indore")
    sh1.cell(25, column=1, value="Kolkata")

    if findMasterLoop != 0:
        for x in range(0, findMasterLoop + 1):
            if x == 0:
                addData = start[1]
            else:
                addData = 1
            month += addData
            if x == 0:
                monthRange = monthrange(start[2], month)
                tempDifference = monthRange[1] - start[0] + 1
                masterLoop += tempDifference

                noOfMonth.append(monthRange[1])
            else:
                if x == findMasterLoop:
                    masterLoop += end[0]
                    noOfMonth.append(monthRange[1])
                else:
                    monthRange = monthrange(start[2], month)
                    tempDifference = monthRange[1] - 0
                    masterLoop += tempDifference
                    noOfMonth.append(monthRange[1])
    else:
        masterLoop = end[0] - start[0]
    indexmonth = 0
    tempdate = 0
    date = 0
    val = 1
    if len(noOfMonth) == 0:
        val = 2
    else:
        val = 1
    for main in range(1, masterLoop + val):
        bo = 0
        if main == 1:
            tempdate = start[0]
            tempmonth = start[1] + indexmonth
            date = "0" + str(tempdate)

            if len(str(tempmonth)) == 1:
                month = "0" + str(tempmonth)
            else:
                month = tempmonth
        else:
            tempmonth = start[1] + indexmonth
            tempdate += 1
            if len(str(tempdate)) == 1:
                date = "0" + str(tempdate)
            else:
                date = tempdate
            if len(str(tempmonth)) == 1:
                month = "0" + str(tempmonth)
            else:
                month = tempmonth

        if len(noOfMonth) != 0:
            try:
                if tempdate == noOfMonth[indexmonth]:
                    if len(str(tempmonth)) == 1:
                        month = "0" + str(tempmonth)
                    else:
                        month = tempmonth
                    date = tempdate

                    tempmonth = 0
                    tempdate = 0
                    indexmonth += 1
                    bo = 1
            except:
                ""

        # print(date)
        # print("--------->", month)
        for row in range(0, len(newData)):
            fetchingcolumn = newData["Pickup Completion Date"][row].date()
            dateOld = str(fetchingcolumn)
            dateModify = ""
            if dateOld != "NaT":
                dateModify = dateOld
            if dateModify == "2021-" + str(month) + "-" + str(date):
                dateActual = fetchingcolumn.strftime("%d-%b-%Y")
                city = newData["City"][row]
                if city == "Delhi":
                    delhiCount += 1
                if city == "Noida":
                    nodiaCount += 1
                if city == "Ghaziabad":
                    gazibadCount += 1
                if city == "Gurgaon":
                    gurgoanCount += 1
                if city == "Faridabad":
                    faridadCount += 1
                if city == "Chandigarh":
                    chatishgarhCount += 1
                if city == "Jaipur":
                    jaipurCount += 1
                if city == "Mumbai" or city == "Navi Mumbai":
                    mumbaiCount += 1
                if city == "Pune":
                    puneCount += 1
                if city == "Ahmedabad":
                    ahmedabadCount += 1
                if city == "Nagpur":
                    nagarCount += 1
                if city == "Surat":
                    suratCount += 1
                if city == "Bengaluru":
                    benguluruCount += 1
                if city == "Hyderabad":
                    hyderabadCount += 1
                if city == "Chennai":
                    chennaiCount += 1
                if city == "Kochi":
                    kochiCount += 1
                if city == "Coimbatore":
                    coimbatoreCount += 1
                if city == "Kanpur":
                    kanpurCount += 1
                if city == "Lucknow":
                    lucknowCount += 1
                if city == "Indore":
                    indoreCount += 1
                if city == "Kolkata":
                    kolkataCount += 1
        sh1.cell(row=1, column=main + 1, value=dateActual)
        sh1.cell(row=2, column=main + 1, value=delhiCount)
        sh1.cell(row=3, column=main + 1, value=nodiaCount)
        sh1.cell(row=4, column=main + 1, value=gazibadCount)
        sh1.cell(row=5, column=main + 1, value=gurgoanCount)
        sh1.cell(row=6, column=main + 1, value=faridadCount)
        sh1.cell(row=7, column=main + 1, value=chatishgarhCount)
        sh1.cell(row=8, column=main + 1, value=jaipurCount)
        exactValue1 = 0
        for x in range(2, 9):
            convertedvalue1 = sh1.cell(x, column=main + 1).value
            exactValue1 += int(convertedvalue1)
        sh1.cell(9, column=main + 1, value=exactValue1)
        sh1.cell(10, column=main + 1, value=mumbaiCount)
        sh1.cell(11, column=main + 1, value=puneCount)
        sh1.cell(12, column=main + 1, value=ahmedabadCount)
        sh1.cell(13, column=main + 1, value=nagarCount)
        sh1.cell(14, column=main + 1, value=suratCount)
        exactValue2 = 0
        for x in range(10, 15):
            convertedvalue2 = sh1.cell(x, column=main + 1).value
            exactValue2 += int(convertedvalue2)
        sh1.cell(15, column=main + 1, value=exactValue2)
        sh1.cell(16, column=main + 1, value=benguluruCount)
        sh1.cell(17, column=main + 1, value=hyderabadCount)
        sh1.cell(18, column=main + 1, value=chennaiCount)
        exactValue3 = 0
        for x in range(16, 19):
            convertedvalue3 = sh1.cell(x, column=main + 1).value
            exactValue3 += int(convertedvalue3)
        sh1.cell(19, column=main + 1, value=exactValue3)
        sh1.cell(20, column=main + 1, value=kochiCount)
        sh1.cell(21, column=main + 1, value=coimbatoreCount)
        sh1.cell(22, column=main + 1, value=kanpurCount)
        sh1.cell(23, column=main + 1, value=lucknowCount)
        sh1.cell(24, column=main + 1, value=indoreCount)
        sh1.cell(25, column=main + 1, value=kolkataCount)
        delhiCount = 0
        nodiaCount = 0
        gazibadCount = 0
        gurgoanCount = 0
        faridadCount = 0
        chatishgarhCount = 0
        jaipurCount = 0
        mumbaiCount = 0
        puneCount = 0
        ahmedabadCount = 0
        nagarCount = 0
        suratCount = 0
        benguluruCount = 0
        hyderabadCount = 0
        chennaiCount = 0
        kochiCount = 0
        coimbatoreCount = 0
        kanpurCount = 0
        lucknowCount = 0
        indoreCount = 0
        kolkataCount = 0


def calulateTotal():
    wb = openpyxl.load_workbook("C:\\Users\\S.S\\PycharmProjects\\locationData\\" + excelName)
    sh1 = wb["Sheet1"]
    totalRow = sh1.max_row
    newValue = 0
    totalColumn = sh1.max_column
    for x in range(1, totalRow + 1):
        for col in range(1, totalColumn + 1):
            rowValue = sh1.cell(row=x, column=col + 1).value
            if rowValue != None:
                try:
                    newValue += int(rowValue)
                except:
                    ""
            if rowValue is None:
                if x == 1:
                    print('weee')
                    sh1.cell(row=x, column=col + 1, value="Pick Up Done")
                else:
                    sh1.cell(row=x, column=col + 1, value=newValue)
                    newValue = 0
    wb.save(excelName)

    wb = openpyxl.load_workbook("C:\\Users\\S.S\\PycharmProjects\\locationData\\" + excelName)
    sh1 = wb["Sheet1"]
    totalColumn = sh1.max_column
    print(totalColumn)
    totalRow = sh1.max_row

    wb2 = openpyxl.load_workbook("C:\\Users\\S.S\\PycharmProjects\\locationData\\pickUpTarget.xlsx")
    sh2 = wb2["Sheet1"]
    totalRow2 = sh2.max_row
    for de in range(1, totalRow2 + 1):
        rowValue2 = sh2.cell(row=de, column=1).value
        if de == 1:
            sh1.merge_cells(start_row=2, start_column=totalColumn + 1, end_row=4, end_column=totalColumn + 1)
            sh1.cell(row=2, column=totalColumn + 1, value=rowValue2)
            wb2.save(excelName)
        if de == 2:
            sh1.merge_cells(start_row=5, start_column=totalColumn + 1, end_row=6, end_column=totalColumn + 1)
            sh1.cell(row=5, column=totalColumn + 1, value=rowValue2)
        if de == 3:
            sh1.cell(row=7, column=totalColumn + 1, value=rowValue2)
        if de == 4:
            sh1.cell(row=8, column=totalColumn + 1, value=rowValue2)
        if de == 5:
            sh1.cell(row=10, column=totalColumn + 1, value=rowValue2)
        if de == 6:
            sh1.cell(row=11, column=totalColumn + 1, value=rowValue2)
        if de == 7:
            sh1.cell(row=12, column=totalColumn + 1, value=rowValue2)

        if de == 8:
            sh1.cell(row=13, column=totalColumn + 1, value=rowValue2)
        if de == 9:
            sh1.cell(row=14, column=totalColumn + 1, value=rowValue2)

        if de == 10:
            sh1.cell(row=16, column=totalColumn + 1, value=rowValue2)

        if de == 11:
            sh1.cell(row=17, column=totalColumn + 1, value=rowValue2)
        if de == 12:
            sh1.cell(row=18, column=totalColumn + 1, value=rowValue2)
        if de == 13:
            sh1.cell(row=20, column=totalColumn + 1, value=rowValue2)

        if de == 14:
            sh1.cell(row=21, column=totalColumn + 1, value=rowValue2)
        if de == 15:
            sh1.cell(row=22, column=totalColumn + 1, value=rowValue2)
        if de == 16:
            sh1.cell(row=23, column=totalColumn + 1, value=rowValue2)
        if de == 17:
            sh1.cell(row=24, column=totalColumn + 1, value=rowValue2)
        if de == 18:
            sh1.cell(row=25, column=totalColumn + 1, value=rowValue2)

    wb.save(excelName)


def overAchived():
    counto = 0
    counto2 = 0
    count = 0
    count2 = 0
    wb = openpyxl.load_workbook("C:\\Users\\S.S\\PycharmProjects\\locationData\\" + excelName)
    sh1 = wb["Sheet1"]
    totalRow = sh1.max_row
    totalColumn = sh1.max_column
    for value in range(1, totalRow + 1):

        if value != 1 and value < 5:
            count += sh1.cell(row=value, column=totalColumn - 1).value
            if sh1.cell(row=value, column=totalColumn).value is not None:
                count2 += sh1.cell(row=value, column=totalColumn).value
            if (value == 4):
                total = (count / count2) * 100
                sh1.cell(row=value, column=totalColumn + 1, value=total)
        elif value >= 5 and value < 7:
            counto += sh1.cell(row=value, column=totalColumn - 1).value

            if sh1.cell(row=value, column=totalColumn).value is not None:
                counto2 += sh1.cell(row=value, column=totalColumn).value
                print('------>', counto)
            if (value == 6):
                total = (counto / counto2) * 100
                sh1.cell(row=value, column=totalColumn + 1, value=total)
        else:
            newcount = 0
            newcountd = 0
            newcount = sh1.cell(row=value, column=totalColumn - 1).value
            newcountd = sh1.cell(row=value, column=totalColumn).value
            if newcountd and newcount != None:
                total = (int(newcount) / int(newcountd)) * 100
                sh1.cell(row=value, column=totalColumn + 1, value=total)
    wb.save(excelName)


def color():
    wb = openpyxl.load_workbook("C:\\Users\\S.S\\PycharmProjects\\locationData\\" + excelName)
    sh1 = wb["Sheet1"]
    totalRow = sh1.max_row
    totalColumn = sh1.max_column
    total = 0
    total2 = 0
    new = ""
    for give in range(1, totalRow + 1):
        if (give == 1):
            sh1.cell(row=give, column=totalColumn - 1, value="Pick Up Target")
        elif (give == 9):
            sh1.cell(row=9, column=totalColumn - 1, value=total)
            total = 0
        if give != 1:
            values = sh1.cell(row=give, column=totalColumn - 1).value
            if give < 9:
                if values != None:
                    total += int(values)
            elif (give > 9):
                if values != None:
                    total += int(values)
                else:
                    values = sh1.cell(row=give, column=totalColumn - 1, value=total)
                    total = 0

        if (give == 1):
            sh1.cell(row=give, column=totalColumn, value="% OverAchieved")
        elif (give == 9):
            sh1.cell(row=9, column=totalColumn, value=total2)
            total2 = 0
        if give != 1:
            values = sh1.cell(row=give, column=totalColumn).value
            if give < 9:
                if values != None:
                    total2 += int(values)
            elif (give > 9):
                if values != None:
                    total2 += int(values)
                else:
                    values = sh1.cell(row=give, column=totalColumn, value=total2)
                    total2 = 0

        if give > 1:
            valz = sh1.cell(row=give, column=totalColumn).value
            if valz != None:
                permanent = int(valz)
                if permanent >= 100:
                    sh1.cell(give, column=totalColumn).fill = PatternFill("solid", fgColor="98FB98")
                else:
                    sh1.cell(give, column=totalColumn).fill = PatternFill("solid", fgColor="F08080")

            new = sh1.cell(row=give, column=1).value
            if new == "Delhi NCR" or new == "West" or new == "South":
                for x in range(1, totalColumn):
                    sh1.cell(give, column=x).fill = PatternFill("solid", fgColor="FFFF00")

    # sh1.cell(22, column=8).fill = PatternFill("solid", fgColor="FFFF00")
    wb.save(excelName)


fetchData([1, 3, 2021], [5, 3, 2021])
wb.save(excelName)
calulateTotal()
overAchived()
color()
